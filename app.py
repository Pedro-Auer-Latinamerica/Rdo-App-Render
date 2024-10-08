from flask import Flask, render_template, request, send_file, make_response
import openpyxl
from openpyxl.drawing.image import Image
from werkzeug.utils import secure_filename
import os
import sqlite3
from PIL import Image as PilImage

app = Flask(__name__)

# Configurações
EXCEL_FILE = "RDO_modificado.xlsx"  # Arquivo Excel original
UPLOAD_FOLDER = "uploads"
TEMP_FOLDER = "temp"
DATABASE = 'database.db'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Função para inicializar o banco de dados
def init_db():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS forms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente TEXT,
            contrato TEXT,
            escopo TEXT,
            inicio TEXT,
            termino TEXT,
            dia TEXT,
            horario_trabalho TEXT,
            atividades TEXT,
            observacoes TEXT,
            foto TEXT,
            foto2 TEXT,
            foto3 TEXT
        )
    ''')
    conn.commit()
    conn.close()

# Função para converter valores numéricos corretamente
def converter_para_numero(valor):
    try:
        return float(valor)
    except (ValueError, TypeError):
        return valor  # Retorna o valor original se não puder converter

# Função para redimensionar a imagem para 6cm x 7cm
def redimensionar_imagem(caminho_imagem, largura_cm=7, altura_cm=6):
    img = PilImage.open(caminho_imagem)
    # Convertendo centímetros para pixels (considerando 96 dpi)
    largura_px = int(largura_cm * 37.8)  # 1cm = 37.8 pixels (96 DPI)
    altura_px = int(altura_cm * 37.8)
    img_resized = img.resize((largura_px, altura_px))
    img_resized.save(caminho_imagem)

# Função para adicionar imagem na célula 
def adicionar_imagem_centralizada(worksheet, caminho_imagem, cell):
    img = Image(caminho_imagem)
    img.anchor = cell  # Definindo a âncora para inserir a imagem na célula especificada
    worksheet.add_image(img)

# Função para atualizar o Excel temporário sem alterar formatação original
def atualizar_excel(dados, arquivo_excel):
    wb = openpyxl.load_workbook(arquivo_excel)
    ws = wb.active

    # Atualizando os campos específicos
    ws['Y3'] = converter_para_numero(dados['folha'])
    ws['C5'] = dados['cliente']
    ws['J5'] = dados['contrato']
    ws['R5'] = dados['escopo']
    ws['D9'] = dados['inicio']
    ws['D10'] = dados['termino']
    ws['R8'] = dados['dia']
    ws['Y14'] = dados['horario_trabalho']
    ws['A27'] = dados['atividades']
    ws['A61'] = dados['observacoes_contratante']
    
    # Inserir imagens no "Registro fotográfico" sem alterar a formatação
    if 'foto' in dados and dados['foto']:
        redimensionar_imagem(dados['foto'], largura_cm=7, altura_cm=6)
        adicionar_imagem_centralizada(ws, dados['foto'], 'A44')

    if 'foto2' in dados and dados['foto2']:
        redimensionar_imagem(dados['foto2'], largura_cm=7, altura_cm=6)
        adicionar_imagem_centralizada(ws, dados['foto2'], 'J44')

    if 'foto3' in dados and dados['foto3']:
        redimensionar_imagem(dados['foto3'], largura_cm=7, altura_cm=6)
        adicionar_imagem_centralizada(ws, dados['foto3'], 'U44')

      # Tratando a seleção de 'Condições de Tempo'

    condicoes_tempo_map = {

        'tempo_bom': 'A13',

        'chuva_leve': 'A14',

        'chuva_forte': 'A15',

        'chuva_fora_do_turno': 'A16'

    }

    if 'tempo' in dados and dados['tempo'] in condicoes_tempo_map:

        for cell in ['A13', 'A14', 'A15', 'A16']:

            ws[cell] = ""

        ws[condicoes_tempo_map[dados['tempo']]] = 'X'

    # Tratando a seleção de 'Acidentes'

    acidentes_map = {

        'NÃO_HOUVE': 'I13',

        'SEM_AFASTAMENTO': 'I14',

        'COM_AFASTAMENTO': 'I15',

        'DANOS_MATERIAIS': 'I16'

    }

    if 'ACIDENTES' in dados and dados['ACIDENTES'] in acidentes_map:

        for cell in ['I13', 'I14', 'I15', 'I16']:

            ws[cell] = ""

        ws[acidentes_map[dados['ACIDENTES']]] = 'X'

    # Tratando a seleção de 'Condições da Área'

    condicoes_area_map = {

        'OPERÁVEL': 'P13',

        'OPERÁVEL_PARCIALMENTE': 'P14',

        'INOPERÁVEL': 'P15'

    }

    if 'Area' in dados and dados['Area'] in condicoes_area_map:

        for cell in ['P13', 'P14', 'P15']:

            ws[cell] = ""

        ws[condicoes_area_map[dados['Area']]] = 'X'

    # Tratando as entradas de M.O Indireta

    for i in range(1, 5):

        ws[f'A{20 + (i - 1)}'] = dados.get(f'mo_indireta_desc_{i}', '')

        ws[f'D{20 + (i - 1)}'] = converter_para_numero(dados.get(f'mo_indireta_pres_{i}', ''))

        ws[f'F{20 + (i - 1)}'] = converter_para_numero(dados.get(f'mo_indireta_fc_{i}', ''))

    # Tratando as entradas de M.O Direta

    for i in range(1, 5):

        ws[f'H{20 + (i - 1)}'] = dados.get(f'mo_direta_desc_{i}', '')

        ws[f'K{20 + (i - 1)}'] = converter_para_numero(dados.get(f'mo_direta_pres_{i}', ''))

        ws[f'M{20 + (i - 1)}'] = converter_para_numero(dados.get(f'mo_direta_fc_{i}', ''))

    # Tratando as entradas de Subempreiteiros

    for i in range(1, 5):

        ws[f'O{20 + (i - 1)}'] = dados.get(f'subempreiteiro_desc_{i}', '')

        ws[f'R{20 + (i - 1)}'] = converter_para_numero(dados.get(f'subempreiteiro_pres_{i}', ''))

        ws[f'T{20 + (i - 1)}'] = converter_para_numero(dados.get(f'subempreiteiro_fc_{i}', ''))

    # Tratando as entradas de Equipamentos

    for i in range(1, 5):

        ws[f'V{20 + (i - 1)}'] = dados.get(f'equipamentos_desc_{i}', '')

        ws[f'Z{20 + (i - 1)}'] = converter_para_numero(dados.get(f'equipamentos_quant_{i}', ''))

    # Criando arquivo temporário
    temp_file = os.path.join(TEMP_FOLDER, f"RDO_{dados['cliente']}.xlsx")
    wb.save(temp_file)
    return temp_file

# Rota para exibir o formulário
@app.route('/')
def index():
    return render_template('index.html')

# Rota para receber os dados do formulário e atualizar o Excel
@app.route('/submit', methods=['POST'])
def submit_form():
    dados = request.form.to_dict()

    # Tratando upload de imagens
    if 'foto' in request.files:
        foto = request.files['foto']
        if foto.filename != '':
            filename = secure_filename(foto.filename)
            foto_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            foto.save(foto_path)
            dados['foto'] = foto_path

    if 'foto2' in request.files:
        foto2 = request.files['foto2']
        if foto2.filename != '':
            filename = secure_filename(foto2.filename)
            foto2_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            foto2.save(foto2_path)
            dados['foto2'] = foto2_path

    if 'foto3' in request.files:
        foto3 = request.files['foto3']
        if foto3.filename != '':
            filename = secure_filename(foto3.filename)
            foto3_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            foto3.save(foto3_path)
            dados['foto3'] = foto3_path

    # Atualizando o arquivo Excel temporário
    temp_file = atualizar_excel(dados, EXCEL_FILE)
    
    # Inserir os dados no banco de dados
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO forms (cliente, contrato, escopo, inicio, termino, dia, horario_trabalho, atividades, observacoes, foto)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (dados['cliente'], dados['contrato'], dados['escopo'], dados['inicio'], dados['termino'], dados['dia'],
          dados['horario_trabalho'], dados['atividades'], dados['observacoes_contratante'], dados.get('foto')))
    conn.commit()
    conn.close()

    # Preparar a resposta com cabeçalhos para desativar o cache e permitir o download
    response = make_response(send_file(temp_file, as_attachment=True))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'

# Após o envio, apagar o arquivo temporário
    @response.call_on_close
    def cleanup():
        if os.path.exists(temp_file):
            os.remove(temp_file)

    return response

if __name__ == '__main__':
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    if not os.path.exists(TEMP_FOLDER):
        os.makedirs(TEMP_FOLDER)
    
    # Inicializar o banco de dados
    init_db()
    
    app.run()
